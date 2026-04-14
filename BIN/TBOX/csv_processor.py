#!/usr/bin/env python3
"""
Скрипт для обработки CSV файлов банковских выписок
Автор: TBOX AI Assistant
Дата: 23.03.2026
"""

import pandas as pd
import os
from datetime import datetime
import sys
import argparse

def parse_arguments():
    """
    Парсит аргументы командной строки
    """
    parser = argparse.ArgumentParser(description='Обработка CSV файлов банковских выписок')
    parser.add_argument('--card', type=str, help='Фильтр по Card Name')
    parser.add_argument('--file', type=str, help='Имя CSV файла для обработки')
    parser.add_argument('--input', type=str, help='Папка с CSV файлами')
    parser.add_argument('--output', type=str, help='Папка для сохранения результата')
    return parser.parse_args()

def calculate_smart_rate(row):
    """Умная логика расчета Rate"""
    if row['Amount'] == 0:
        return 0
    
    # Исключения валют - всегда оставляем Foreign Exchange Rate как есть
    excluded_currencies = ['KWD', 'BHD', 'OMR', 'JOD', 'CHF']
    foreign_currency = row.get('Foreign Currency', '')
    
    # Получаем существующий курс
    existing_rate = row.get('Foreign Exchange Rate', 0)
    
    # Если валюта-исключение или курс > 1, оставляем как есть
    if foreign_currency in excluded_currencies or existing_rate > 1:
        return existing_rate
    
    # Иначе делаем расчет: -Foreign Amount / Amount
    if row['Amount'] != 0:
        return abs(row['Foreign Amount'] / row['Amount'])
    
    return 0

def process_csv_file(csv_file_path, card_name_filter=None):
    """
    Обрабатывает CSV файл и создает Excel отчет
    """
    
    try:
        # Читаем CSV файл
        df = pd.read_csv(csv_file_path)
        total_records = len(df)
        pages_count = len(df['Card Name'].unique())
        # 3. Фильтр по Card Name
        if card_name_filter and 'Card Name' in df.columns:
            original_count = len(df)
            df = df[df['Card Name'] == card_name_filter]
        
        # 4. Фильтр по Status - теперь все статусы
        if 'Status' in df.columns:
            original_count = len(df)
            # Не фильтруем - берем все статусы
        else:
            print(f"⚠️ Колонка 'Status' не найдена")
        
        # 5. Сортировка по Date (UTC) - старые → новые
        date_columns = [col for col in df.columns if 'Date' in col]
        if date_columns:
            date_col = date_columns[0]  # Берем первую колонку с Date
            
            # Конвертируем дату если нужно
            if df[date_col].dtype == 'object':
                try:
                    df[date_col] = pd.to_datetime(df[date_col])
                except:
                    print(f"⚠️ Не удалось конвертировать дату, оставляем как есть")
            
            df = df.sort_values(date_col, ascending=True)  # старые → новые
        else:
            print(f"⚠️ Колонка с датой не найдена")
        
        # 6. Оставляем нужные поля
        required_fields = [
            'Date (UTC)', 'Description', 'Amount', 
            'Foreign Amount', 'Foreign Currency', 'Foreign Exchange Rate', 'Status', 'Card Name'
        ]
        
        available_fields = [field for field in required_fields if field in df.columns]
        missing_fields = [field for field in required_fields if field not in df.columns]
        
        if missing_fields:
            print(f"⚠️ Отсутствуют поля: {missing_fields}")
        
        if available_fields:
            df_result = df[available_fields].copy()
        else:
            print(f"❌ Нет доступных полей для обработки")
            return None
        
        # 7. Добавляем поле Rate с умной логикой
        if 'Foreign Amount' in df_result.columns and 'Amount' in df_result.columns:
            # Избегаем деления на ноль
            df_result['Rate'] = df_result.apply(
                lambda row: calculate_smart_rate(row), axis=1
            )
          
        
        # Создаем Excel файл с отдельными листами для каждой карты
        output_file = csv_file_path.replace('.csv', '_processed.xlsx')
        
        # Получаем все уникальные Card Name
        all_cards = df['Card Name'].unique()
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Обрабатываем каждую карту отдельно
            for card_name in all_cards:
                # Фильтруем по карте (все статусы)
                card_data = df[df['Card Name'] == card_name].copy()
                
                # Сортируем по дате
                if 'Date (UTC)' in card_data.columns:
                    card_data = card_data.sort_values('Date (UTC)', ascending=True)
                
                # Оставляем нужные поля
                card_result = card_data[available_fields].copy()
                
                # Добавляем поле Rate
                if 'Foreign Amount' in card_result.columns and 'Amount' in card_result.columns:
                    card_result['Rate'] = card_result.apply(
                        lambda row: calculate_smart_rate(row), axis=1
                    )
                
                # Создаем понятное имя листа
                sheet_name = f"{card_name}"  # Emp 1 -> Card_1
                if len(sheet_name) > 31:  # Excel ограничение
                    sheet_name = f"Card_{card_name[:10]}"
                
                # Сохраняем на отдельный лист
                card_result.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Применяем красную подсветку для не-settled статусов
                worksheet = writer.sheets[sheet_name]
                
                # Находим колонку Status
                status_col = None
                for col_idx, col_name in enumerate(card_result.columns, 1):
                    if col_name == 'Status':
                        status_col = col_idx
                        break
                
                # Применяем красную заливку для не-settled
                if status_col:
                    from openpyxl.styles import PatternFill
                    red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
                    
                    for row in range(2, len(card_result) + 2):  # +2 для заголовка
                        cell_value = worksheet.cell(row=row, column=status_col).value
                        if cell_value and cell_value.lower() != 'settled':
                            # Закрашиваем всю строку красным
                            for col in range(1, len(card_result.columns) + 1):
                                worksheet.cell(row=row, column=col).fill = red_fill
                
                # Считаем статистику для карты
                card_amount = card_result['Amount'].sum() if 'Amount' in card_result.columns else 0
                card_count = len(card_result)
                settled_count = len(card_result[card_result['Status'].str.lower() == 'settled']) if 'Status' in card_result.columns else 0
                
                # Считаем сумму только settled
                settled_amount = 0
                if 'Status' in card_result.columns and 'Amount' in card_result.columns:
                    settled_data = card_result[card_result['Status'].str.lower() == 'settled']
                    settled_amount = settled_data['Amount'].sum()
                
                print(f"📄 Лист '{sheet_name:<8}': {card_count:>3} записей ({settled_count:>3} settled), сумма: {card_amount:>10,.2f} ({settled_amount:>10,.2f})")
        # 8. Считаем общую сумму по Amount
        if 'Amount' in df_result.columns:
            total_amount = df_result['Amount'].sum()
        else:
            total_amount = 0
            print(f"⚠️ Поле Amount не найдено")
        print(f"📊 Загружено записей: {total_records} | 📑 Создано страниц: {pages_count} | 💰 Общая сумма по Amount: {total_amount:,.2f}")        
        return output_file
        
    except Exception as e:
        print(f"❌ Ошибка обработки: {e}")
        return None

def main():
    """
    Главная функция
    """
    # Парсим аргументы
    args = parse_arguments()
    
    print("=" * 50)
    
    # Путь к папке 01_INBOX (или из параметра)
    if args.input:
        inbox_path = args.input
    else:
        inbox_path = "/Users/kivaonmac/Documents/AI_Lab/01_INBOX"
    
    print(f"📁 Папка: {inbox_path}")
    
    # Ищем CSV файлы
    csv_files = [f for f in os.listdir(inbox_path) if f.endswith('.csv')]
    
    if not csv_files:
        print(f"❌ CSV файлы не найдены")
        return
    
    print(f"📁 Найдены файлы:")
    for i, file in enumerate(csv_files, 1):
        print(f"   {i} - {file}")
    
    # Выбираем файл
    if args.file:
        csv_file = args.file
    else:
        try:
            choice = input(f"📁 Выберите файл: ").strip()
            if choice:
                choice_num = int(choice)
                if 1 <= choice_num <= len(csv_files):
                    csv_file = csv_files[choice_num - 1]
                else:
                    csv_file = csv_files[0]
            else:
                csv_file = csv_files[0]  # Take newest file
        except ValueError:
            print(f"⚠️ Неверный ввод, берем первый")
            csv_file = csv_files[0]
    
    print(f"🔄 Обработка файла csv: {csv_file}")
    csv_path = os.path.join(inbox_path, csv_file)
    
    # Получаем доступные карты
    try:
        df_full = pd.read_csv(csv_path)  # Читаем весь файл
        if 'Card Name' in df_full.columns:
            unique_cards = df_full['Card Name'].dropna().unique().tolist()
            print(f"💳 Available cards:")
            for i, card in enumerate(unique_cards, 1):
                print(f"   {i} - {card}")
            
            # Фильтр по умолчанию или выбор
            if args.card:
                card_filter = args.card
                print(f"🎯 Фильтр карты: {card_filter} (указан)")
            else:
                choice = input(f"🎯 Выберите карту (или Enter для всех): ").strip()
                if choice:
                    try:
                        choice_num = int(choice)
                        if 1 <= choice_num <= len(unique_cards):
                            card_filter = unique_cards[choice_num - 1]
                            print(f"✅ Выбрана: {card_filter}")
                        else:
                            print(f"⚠️ Неверный выбор, обрабатываем все")
                            card_filter = None
                    except ValueError:
                        print(f"⚠️ Неверный ввод, обрабатываем все")
                        card_filter = None
                else:
                    card_filter = None
                    print(f"все карты")
        else:
            print(f"⚠️ Колонка 'Card Name' не найдена")
            card_filter = None
    except Exception as e:
        print(f"⚠️ Не удалось прочитать файл: {e}")
        card_filter = None
    
    # Обрабатываем файл
    result_file = process_csv_file(csv_path, card_filter)
    
    # Папка для сохранения
    if args.output:
        output_dir = args.output
        if result_file and os.path.exists(result_file):
            old_name = os.path.basename(result_file)
            new_path = os.path.join(output_dir, old_name)
            os.rename(result_file, new_path)
            result_file = new_path
    
    if result_file:
        print(f"✅ Обработка окончена -> результат записан в: {os.path.basename(result_file)}")
    else:
        print(f"❌ ОБРАБОТКА НЕ УДАЛАСЬ!")
if __name__ == "__main__":
    main()
